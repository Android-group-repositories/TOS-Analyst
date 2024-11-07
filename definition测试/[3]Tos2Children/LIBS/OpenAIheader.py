

from openpyxl import load_workbook
def get_Number_of_tos():
    workbook = load_workbook(filename='inputSentence1.xlsx')

# 选择活动的工作表
    sheet = workbook.active

    # 获取列数
    num_columns = sheet.max_row
    print(num_columns)
    return num_columns - 1

# KIMI
API_KEY = "sk-6oUHGgSeXProgGtghTwo25ZJfFsQ2m674zetmmU4WufYB6YI"
BASE_URL = "https://api.moonshot.cn/v1"
MODEL= "moonshot-v1-8k"


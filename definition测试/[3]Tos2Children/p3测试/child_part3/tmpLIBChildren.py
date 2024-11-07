P3_promptType1 = r"""
For type 1, there are 4 basic patterns for rewriting sentence structures:
1. your app can BOOL VERB_PRESENT DATA ; USER ; Even obtain user consent
2. your app can BOOL VERB_PRESENT DATA ; USER ; when CONDITION
3. your app can BOOL VERB_PRESENT SERVICE ; USER ; Even obtain user consent
4. your app can BOOL VERB_PRESENT SERVICE ; USER ; when CONDITION

5. your app NEED VERB_PRESENT DATA ; USER ; when CONDITION          
6. your app can BOOL VERB_PRESENT DATA or SERVICE  ;  to PURPOSE  ;  Even obtain user consent


for SERVICE:
首字母大写 Services
带the的The service
带第一人称的our service
都有的The Service
可以统一归类为 sdk-service


Please study the following examples; which follow:
-<text>	 You must not collect personal information from minors under any circumstances "
-pattern 1	your app can BOOL VERB_PRESENT DATA ; USER ; Even obtain user consent
-Output	your app {cannot} {collect} {personal information}; {user:minors} ; Even obtain user consent
When the Company plans to gather; utilize; or disclose the personal location data of children younger than 14; it must first secure the approval of both the child and their legal guardian.

-<text>	  : When the Company plans to gather; utilize; or disclose the personal location data of children younger than 14; it must first secure the approval of both the child and their legal guardian.
-pattern 2:	your app can BOOL VERB_PRESENT DATA ; USER ; when CONDITION
-Output	your app {cannot} {collect;use;disclose} {personal location data}; {user: under 14} ; when {secure the approval of both the child and guardian}

-<text>	  : The Service is not intended for and should not be used by anyone under the age of 16. 
-pattern 3:	your app can BOOL VERB_PRESENT SERVICE ; Even obtain user consent 
-Output	your app {cannot} {use sdk-servic} ; {user: under the age of 16. } ;Even obtain user consent


-<text>	    : To use our services and be bound by COPPA; CCPA; or GDPR; you must have obtained verifiable and voluntary consent from the child's parents or guardians.
-pattern 4  :	your app can BOOL VERB_PRESENT SERVICE  ; when CONDITION  
-Output	    :your app {can} {use sdk-service} ;{user:child} ;  when {obtained consent from  child's parents or guardians}


-<text>     :The customer is responsible for maintaining the security of user IDs, passwords, and other access credentials.
-pattern 5  :your app {need} {maintaining security of} { user IDs, passwords, access credentials}  ;   Even obtain user consent

-<text>     :The customer must ensure that these access credentials are kept strictly confidential.
-pattern 5  :your app {need} {kept security of} {  access credentials}  ;   Even obtain user consent


对于 下列 tos, 文中出现了 数据和对数据具体项的解释 两者都需要提取           
-<text>     :You may not collect, capture, use or store PayPal Prohibited User Information including identifiers used for tax or claiming government benefits, national identity numbers, passport numbers 
-pattern 1  :your app {not} {collect, capture, use , store} {PayPal Prohibited User Information INCLUDEING identifiers used for tax  , national identity numbers, passport numbers }  ; {user} ;Even obtain user consent

-<text>     :You may not collect, capture, use or store PayPal Prohibited User Information including identifiers used for tax or claiming government benefits, national identity numbers, passport numbers 
-pattern 1  :your app {not} {collect, capture, use , store} {PayPal Prohibited User Information INCLUDEING identifiers used for tax  , national identity numbers, passport numbers }   ;  {user} ;Even obtain user consent

如果提到了获取信息的目的,请使用模式6

-<text>     :You are not allowed to use authorized materials to obtain or attempt to obtain non-public information of X user.
-pattern 6  :your app {cannot} {use authorized materials} ; { to obtain or attempt to obtain non-public information} ; Even obtain user consent


The output only includes the part after "-Output" and pattern
be like:
"#Type1-pattern2# your app {can} {use service} ;{user:under the age of 16. } ;  Even obtain user consent   "

接下来我会输入一些<text>,按用适合的pattern对句子进行改写
"""

P3_promptType2 = r"""
For type 2; there are 2 basic patterns for rewriting sentence structures:
ENT in type 2 is a third party entity
1. your app can BOOL VERB_PRESENT DATA to third-party ENT , USER , Even obtain user consent;
2. your app can BOOL VERB_PRESENT DATA to ENT , USER , when CONDITION ;



for SERVICE:
首字母大写 Services
带the的The service
带第一人称的our service
都有的The Service
可以统一归类为 sdk-service


Please study the following examples; which follow:
-<text>	    :Developers are not allowed to intentionally share data of users under the age of 16 with Pollfish
-pattern 1	:your app can BOOL VERB_PRESENT DATA to ENT , USER , Even obtain user consent; 
-Output	your app {can not} {share} {user data} to {Pollfish} , {user:Under 16 years old} , Even obtain user consent,

-<text>	    :You are not allowed to share personal identifiers such as IP addresses, IDFA/IDFV, Android ID, Google Play Advertising ID, Google Play store referrer, WindowsHardware ID, Windows NetworkID, Windows Phone device ID, and UUIDs with Adjust partners or other third parties
-pattern 1	:your app can BOOL VERB_PRESENT DATA to ENT , USER , Even obtain user consent
-Output	"your app {can not} {not} {share} {IP addresses, IDFA/IDFV, Android ID, Google Play Advertising ID, Google Play store referrer, WindowsHardware ID, Windows NetworkID, Windows Phone device ID, and UUIDs} to {Adjust , third parties} , {user:Under 16 years old} , Even obtain user consent


-<text>     :If your applications are targeted at children under COPPA, CCPA, CPRA, or similar US privacy laws, you must obtain verifiable user or parental consent for collecting and sharing data with InMobi.
-pattern 2	:your app can BOOL VERB_PRESENT DATA to ENT , when CONDITION 
-Output	your app {can} {collect , share} {Data} to {InMobi}  , {user : children under COPPA, CCPA, CPRA, or similar US privacy laws,} ,when {obtain user or parental consent} 



The output only includes the part after "-Output" and pattern
be like:
"#Type1-pattern2# your app {can} {use service} ;{user:under the age of 16. } ;  Even obtain user consent "

接下来我会输入一些<text>,按用适合的pattern对句子进行改写
"""
# 1. your app can BOOL VERB_PRESENT DATA ; USER ; Even obtain user consent ; OPTOUT
# 2. your app can BOOL VERB_PRESENT DATA ; USER ; when CONDITION ; OPTOUT
# 3. your app can BOOL VERB_PRESENT SERVICE ; USER ; Even obtain user consent ; OPTOUT
# 4. your app can BOOL VERB_PRESENT SERVICE ; USER ; when CONDITION  ; OPTOUT

P3_promptType3 = r"""
For type 2, there is a basic patterns for rewriting sentence structures:

1. privacy police need {....}  ,  include {....}  ,  to opt-out from  {....} 

{}的内容可以是：
- {provide Optout mechanism}
- {TO opt-out : 对象},对象可以是 advertising,information saleing 等

Provide children guardian the ability to opt-out from the sale of their information



Please study the following examples; which follow:
-<text>	    :your privacy police  need to describe the opt-out information includes online behavioral or mobile cross-app advertising.
-pattern 2	:privacy police need {to describe the opt-out information}  ,  to opt-out from {online behavioral or mobile cross-app advertising}

-<text>	    :Your privacy policy must include an opt-out link directing users to an industry-recognized opt-out 
-pattern 3	:privacy police need provide a {link to opt-out}  ,  include {an opt-out link}

-<text>	    :Your privacy policy must include an opt-out link directing users to an industry-recognized opt-out from online behavioral advertising.
-pattern 4  :privacy police need provide a {link to opt-out}  ,  include {a conspicuous link}  ,  to opt-out from  {online behavioral advertising} 

Provide children guardian the ability to opt-out from the sale of their information



The output only includes the part after "-Output" and pattern
be like:
"#Type1-pattern4# privacy police need provide a  {link to opt-out}  ,  include {a conspicuous link}  ,  to opt-out from  {online behavioral advertising} "

接下来我会输入一些<text>,按用适合的pattern对句子进行改写
"""


from openpyxl import load_workbook
def get_Number_of_tos():
    workbook = load_workbook(filename='inputSentence1.xlsx')

# 选择活动的工作表
    sheet = workbook.active

    # 获取列数
    num_columns = sheet.max_row
    print(num_columns)
    return num_columns - 1


# KIMI
API_KEY = "sk-kk5QnQbIkzxsASlPA11b444dFfCd467989Bb120b4d8e0aA9"
BASE_URL = "https://api.holdai.top/v1"
MODEL= "moonshot-v1-8k"

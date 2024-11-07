
# （括号里面是注释）
P3_promptType1 = r"""
For type 1, there are 4 basic patterns for rewriting sentence structures:
1. Your privacy policy needs to describe YOU VERB_PRESENT DATA  ;  Even obtain user consent
2. Your privacy policy needs to describe YOU USE SERVICE  ;  Even obtain user consent
3. Your privacy policy needs to describe YOU VERB_PRESENT DATA or SERVICE  ;  for PURPOSE  ;  Even obtain user consent
4. Your privacy policy needs to describe YOU VERB_PRESENT DATA to third-party ENT ; USER ; Even obtain user consent;

when no data or service mentioned  ,  use pattern 5 
5. You provide privacy policy and CONDITION

for SERVICE:表示对服务的使用


Please study the following examples; which follow:
-<text>     :You must develop a privacy policy that clearly and accurately informs end users about the personal data you collect 
Type1-pattern 1  :Your privacy policy needs to describe {you} {collect} {personal data} ; Even obtain user consent


-<text>     :privacy policy provide notice of  use of technology that collects, uses, shares and stores data about end users of your Application 
Type1-pattern 1  :Your privacy policy needs to describe {you} {collects, uses, shares , stores} {data about end users} ; Even obtain user consent

-<text>     :privacy policy provide notice of  use of a tracking pixel, agent or any other visitor identification technology 
-pattern 2  :Your privacy policy needs to describe {you} {use} {tracking pixel, agent or any other visitor identification technology} ; Even obtain user consent

pattern 3 具有使用服务、数据的目的
-<text>     :You must post a Privacy Policy that provides notice of Your use of cookies, mobile device identifiers, or similar technology used to collect data.
Type1-pattern 3  :Your privacy policy needs to describe {you} {use cookies, mobile device identifiers, or similar technology} ; {to collect data} ; Even obtain user consent



-<text>     :The privacy policy must specifically address the use of personal information for behaviorally targeted online advertising.
Type1-pattern 3  :Your privacy policy needs to describe {you} {use} {personal informatio} ; for {behaviorally targeted online advertising}  ;  Even obtain user consent

when no data or service mentioned  ,  use pattern 4 
-<text>     :You must provide a legally adequate privacy notice to End Users about it
Type1-pattern 4  :You provide privacy policy and {legally adequate it to End Users}

-<text>     :You must clearly and conspicuously post a privacy policy within your application, in any store, and on any website where the application may be acquired
Type1-pattern 4  :You provide privacy policy and {post it in store and website}

-<text>     :Licensee shall publish privacy policies and disclosures for the Properties that comply with applicable law and the terms of this EULA
Type1-pattern 4  :You provide privacy policy and {disclosures for the Properties that comply with applicable law and the terms of this EULA}

The output only includes the part after "-Output" and pattern
be like:
"#Type1-pattern2# Your privacy policy needs to describe {you} {collect} {personal data} ; Even obtain user consent"

接下来我会输入一些<text>,按用适合的pattern对句子进行改写
"""

P3_promptType2 = r"""
For type 2; there are 2 basic patterns for rewriting sentence structures:
ENT in type 2 is a third party entity
1. Your privacy policy needs to describe YOU VERB_PRESENT DATA ; to third-party ENT  ;  Even obtain user consent
1. Your privacy policy needs to describe YOU VERB_PRESENT DATA ; to third-party ENT  ;  when CONDITION




for SERVICE:
首字母大写 Services
带the的The service
带第一人称的our service
都有的The Service
可以统一归类为 sdk-service


Please study the following examples; which follow:
-<text>     :Your privacy policy must describe how you and third parties use, collect, and share mobile device advertising identifiers and other personal information
Type2-pattern 1	:Your privacy policy needs to describe YOU {use, collect,  share} {device advertising identifiers , other personal information}  ; to {third parties} ;  Even obtain user consent


-<text>	    :The privacy policy must explain how and why you and your Application use, process, and share user information with Snap and other third parties
Type2-pattern 1	:Your privacy policy needs to describe YOU {use, process, share} { user information}  ; to {Snap , other third parties} ;  Even obtain user consent


The output only includes the part after "-Output" and pattern
be like:
"#Type2-pattern1# Your privacy policy needs to describe YOU {use, collect,  share} {device advertising identifiers , other personal information} to {third parties} ;  Even obtain user consent"

接下来我会输入一些<text>,按用适合的pattern对句子进行改写
"""
# 1. your app can BOOL VERB_PRESENT DATA ; USER ; Even obtain user consent ; OPTOUT
# 2. your app can BOOL VERB_PRESENT DATA ; USER ; when CONDITION ; OPTOUT
# 3. your app can BOOL VERB_PRESENT SERVICE ; USER ; Even obtain user consent ; OPTOUT
# 4. your app can BOOL VERB_PRESENT SERVICE ; USER ; when CONDITION  ; OPTOUT


# # zanshi
# P3_promptType3 = r"""
# For type 2, there are 4 basic patterns for rewriting sentence structures:

# 1. your app can BOOL VERB_PRESENT DATA ; USER ; when CONDITION ; OPTOUT
# 2. your app can BOOL VERB_PRESENT SERVICE ; USER ; when CONDITION  ; OPTOUT

# OPTOUT可以是:
# - {provide Optout mechanism}
# - {TO opt-out : 对象},对象可以是 advertising,information saleing 等

# Provide children guardian the ability to opt-out from the sale of their information

# for SERVICE:
# 首字母大写 Services
# 带the的The service
# 带第一人称的our service
# 都有的The Service
# 可以统一归类为 sdk-service


# Please study the following examples; which follow:
# -<text>	    :You must not collect personal information from minors under any circumstances , the opt-out information includes online behavioral or mobile cross-app advertising
# -pattern 2	:your app can BOOL VERB_PRESENT DATA ; USER ; Even obtain user consent
# -Output	your app {cannot} {collect} {personal information}; {user:minors} ; Even obtain user consent ; {TO opt-out :advertising}

# Provide children guardian the ability to opt-out from the sale of their information



# The output only includes the part after "-Output" and pattern
# be like:
# "#pattern2# your app {can} {use service} ;{user:under the age of 16. } ;  Even obtain user consent  ; TO opt-out : {advertising}  "

# 接下来我会输入一些<text>,按用适合的pattern对句子进行改写
# """


from openpyxl import load_workbook
def get_Number_of_tos():
    workbook = load_workbook(filename='inputSentence1.xlsx')

# 选择活动的工作表
    sheet = workbook.active

    # 获取列数
    num_columns = sheet.max_row
    print(num_columns)
    return num_columns - 1

# KIMI
API_KEY = "sk-kk5QnQbIkzxsASlPA11b444dFfCd467989Bb120b4d8e0aA9"
BASE_URL = "https://api.holdai.top/v1"
MODEL= "moonshot-v1-8k"

